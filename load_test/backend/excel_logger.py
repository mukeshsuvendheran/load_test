from openpyxl import Workbook, load_workbook
import os
from config import EXCEL_FILE

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "RPM", "Voltage", "Current", "Temperature"])
        wb.save(EXCEL_FILE)

def log_data(data):
    wb = load_workbook(EXCEL_FILE) # Load existing workbook
    ws = wb.active

    ws.append([
        data["timestamp"],
        data["rpm"],
        data["voltage"],
        data["current"],
        data["temperature"]
    ])

    wb.save(EXCEL_FILE)